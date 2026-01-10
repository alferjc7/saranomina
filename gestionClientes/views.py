from datetime import datetime
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import CreateView, DeleteView
from gestionClientes.models import t_cliente, t_empresa, UsuarioEmpresa
from gestionClientes.forms import t_cliente_form, t_empresa_form, usuarioempresa_form
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class clienteCreateView(LoginRequiredMixin,CreateView):
    model = t_cliente
    form_class = t_cliente_form
    template_name = 'clientes.html'
    context_object_name = 'clientes'
    success_url = reverse_lazy('clientes')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Clientes'

            ws.append(['NOMBRE', 'CELULAR', 'ESTADO'])

            registros = t_cliente.objects.all()

            for r in registros:
                ws.append([r.nombre_cliente, r.celular, 'ACTIVO' if r.estado_cliente else 'INACTIVO'])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Clientes.xlsx'
            wb.save(response)
            return response 

        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_cliente.objects.all().order_by('-pk')[:30]
        context['registros'] = registros
        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            cliente = t_cliente.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(cliente, field, value)
            cliente.save()
            messages.success(self.request, 'Registro actualizado correctamente')
        else:
            response = super().form_valid(form)
            messages.success(self.request, 'Registro creado correctamente')
        return redirect(self.success_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)

class clienteDeleteView(LoginRequiredMixin,DeleteView):
    model = t_cliente
    success_url = reverse_lazy('clientes')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Cliente eliminado correctamente')
        return super().post(request, *args, **kwargs)

    

class empresasCreateView(LoginRequiredMixin, CreateView):
    model = t_empresa
    form_class = t_empresa_form
    template_name = 'empresas.html'

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Empresas_cliente'

            ws.append(['CLIENTE', 'CODIGO EMPRESA','NIT EMPRESA','DIGITO VERIFICACION','RAZON SOCIAL'
                       ,'DIRECCION','TELEFONO'])

            cliente = get_object_or_404(t_cliente,pk=self.kwargs['id'])
        
            registros = t_empresa.objects.filter(codigo_cliente=cliente)

            for r in registros:
                ws.append([r.codigo_cliente.nombre_cliente, r.codigo_empresa, r.nit_empresa, 
                           r.digito_verificacion, r.razon_social, r.direccion, r.telefono])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Empresas_cliente.xlsx'
            wb.save(response)
            return response 
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = get_object_or_404(t_cliente,pk=self.kwargs['id'])
        
        context['cliente'] = cliente
        context['registros'] = t_empresa.objects.filter(codigo_cliente=cliente)
        
        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            empresa = t_empresa.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(empresa, field, value)
            empresa.save()
            messages.success(self.request, 'Empresa actualizada correctamente')
        else:
            cliente = get_object_or_404(t_cliente,pk=self.kwargs['id'])
            form.instance.codigo_cliente = cliente
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())
   
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'empresas',
            kwargs={'id': self.kwargs['id']}
        )
    
class empresasDeleteView(LoginRequiredMixin,DeleteView):
    model = t_empresa
    login_url = 'accounts/login'
    
    def get_success_url(self):
        return reverse_lazy(
            'empresas',
            kwargs={'id': self.object.codigo_cliente.id}
        )
    
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)
    
class userempresaCreateView(LoginRequiredMixin, CreateView):
    model = UsuarioEmpresa
    form_class = usuarioempresa_form
    template_name = 'user_empresa.html'
    context_object_name = 'user_empresa'
    success_url = reverse_lazy('user_empresa')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Asignaciones_empresa'

            ws.append(['USUARIO', 'EMPRESA', 'ESTADO', 'FECHA ASIGNACION'])

            registros = UsuarioEmpresa.objects.all()

            for r in registros:
                ws.append([r.usuario.username, r.empresa.razon_social, 'ACTIVO' if r.activo else 'INACTIVO',r.fecha_asignacion.replace(tzinfo=None) if r.fecha_asignacion else None])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Asignaciones_Empresa.xlsx'
            wb.save(response)
            return response 

        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = UsuarioEmpresa.objects.all().order_by('-pk')[:30]
        context['registros'] = registros
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Registro creado correctamente')
        return super().form_valid(form)
        
    def form_invalid(self, form):
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)

class userempresasDeleteView(LoginRequiredMixin,DeleteView):
    model = UsuarioEmpresa
    success_url = reverse_lazy('user_empresa')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Empresa eliminado correctamente')
        return super().post(request, *args, **kwargs)
    

@login_required
def seleccionar_empresa(request):
    empresas = UsuarioEmpresa.objects.filter(
        usuario=request.user,
        activo=True
    ).select_related('empresa')

    if not empresas.exists():
        return render(request, 'gestionClientes/sin_empresas.html')

    if request.method == 'POST':
        empresa_id = request.POST.get('empresa_id')
        empresa_usuario = empresas.filter(empresa_id=empresa_id).first()
        empresa = empresa_usuario.empresa


        if empresas.filter(empresa_id=empresa_id).exists():
            request.session['empresa_id'] = empresa_id
            request.session['codigo_empresa'] = empresa.codigo_empresa
            request.session['razon_social'] = empresa.razon_social

            return redirect('inicio')  # tu home real

    return render(request, 'seleccionar_emp.html', {
        'empresas': empresas
    })