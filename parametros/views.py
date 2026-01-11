from datetime import datetime
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from parametros.models import (t_tipo_contrato, t_tipo_salario, 
                               t_tipo_cotizante, t_subtipo_cotizante,
                               t_banco, t_entidadesss, t_conceptos_salario, t_tipo_nomina,
                               ParametroGeneral, ParametroDetalle)
from parametros.forms import (tipo_contratoform, tipo_salarioform, 
                              tipo_cotizanteform, subtipo_cotizanteform,
                              bancoform, t_entidadesssform, CargaExcelForm, 
                              t_conceptos_salarioform, t_tipo_nominaform, 
                              t_parametrogeneralform, t_parametrodetllleform)
from django.views.generic import CreateView, DeleteView
from openpyxl import load_workbook
from django.http import HttpResponse
from openpyxl import Workbook


# Create your views here.
class tipos_contratosCreateView(LoginRequiredMixin,CreateView):
    model = t_tipo_contrato
    form_class = tipo_contratoform
    template_name = 'tipos_contratos.html'
    context_object_name = 'tipos_contratos'
    success_url = reverse_lazy('tipos_contratos')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_tipo_contrato.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        tipo = self.request.GET.get('tipo')

        if tipo:
            context['registros'] = t_tipo_contrato.objects.filter(contrato=tipo)



        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            tipo_contrato = t_tipo_contrato.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(tipo_contrato, field, value)
            tipo_contrato.save()
            messages.success(self.request, 'Tipo de contrato actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        
        return redirect(self.success_url)
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)

class tipo_contratosDeleteView(LoginRequiredMixin,DeleteView):
    model = t_tipo_contrato
    success_url = reverse_lazy('tipos_contratos')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de contraro eliminado correctamente')
        return super().post(request, *args, **kwargs)

class tipos_salariosCreateView(LoginRequiredMixin,CreateView):
    model = t_tipo_salario
    form_class = tipo_salarioform
    template_name = 'tipos_salarios.html'
    context_object_name = 'tipos_salarios'
    success_url = reverse_lazy('tipos_salarios')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_tipo_salario.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        tipo = self.request.GET.get('tipo')
        
        estado = self.request.GET.get('estado')


        if tipo:
            context['registros'] = t_tipo_salario.objects.filter(salario=tipo)

        
        if estado:
            context['registros'] = t_tipo_salario.objects.filter(estado=estado)

        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            tipo_salario = t_tipo_salario.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(tipo_salario, field, value)
            tipo_salario.save()
            messages.success(self.request, 'Tipo de salario actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.success_url)

    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class tipo_salariosDeleteView(LoginRequiredMixin,DeleteView):
    model = t_tipo_salario
    success_url = reverse_lazy('tipos_salarios')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de contraro eliminado correctamente')
        return super().post(request, *args, **kwargs)
    

class tipos_cotizantesCreateView(LoginRequiredMixin,CreateView):
    model = t_tipo_cotizante
    form_class = tipo_cotizanteform
    template_name = 'tipos_cotizantes.html'
    context_object_name = 'tipos_cotizantes'
    success_url = reverse_lazy('tipos_cotizantes')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_tipo_cotizante.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        codigo = self.request.GET.get('codigo')
        
        estado = self.request.GET.get('estado')


        if codigo:
            context['registros'] = t_tipo_cotizante.objects.filter(codigo=codigo)

        
        if estado:
            context['registros'] = t_tipo_cotizante.objects.filter(estado=estado)


        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            tipo_cotizante = t_tipo_cotizante.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(tipo_cotizante, field, value)
            tipo_cotizante.save()
            messages.success(self.request, 'Tipo de cotizante actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.success_url)

    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class tipo_cotizantesDeleteView(LoginRequiredMixin,DeleteView):
    model = t_tipo_cotizante
    success_url = reverse_lazy('tipos_cotizantes')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de cotizante eliminado correctamente')
        return super().post(request, *args, **kwargs)


class subtipos_cotizantesCreateView(LoginRequiredMixin,CreateView):
    model = t_tipo_cotizante
    form_class = subtipo_cotizanteform
    template_name = 'subtipos_cotizantes.html'
    context_object_name = 'subtipos_cotizantes'
    success_url = reverse_lazy('subtipos_cotizantes')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_subtipo_cotizante.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        cotizante = self.request.GET.get('cotizante')

        codigo = self.request.GET.get('codigo')
        
        estado = self.request.GET.get('estado')

        
        if cotizante:
            context['registros'] = t_subtipo_cotizante.objects.filter(codigo_cotizante=cotizante)


        if codigo:
            context['registros'] = t_subtipo_cotizante.objects.filter(codigo=codigo)

        
        if estado:
            context['registros'] = t_subtipo_cotizante.objects.filter(estado=estado)


        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            subtipo_cotizante = t_subtipo_cotizante.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(subtipo_cotizante, field, value)
            subtipo_cotizante.save()
            messages.success(self.request, 'Subtipo de cotizante actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        
        return redirect(self.success_url)

    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class subtipo_cotizantesDeleteView(LoginRequiredMixin,DeleteView):
    model = t_subtipo_cotizante
    success_url = reverse_lazy('subtipos_cotizantes')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Subtipo de cotizante eliminado correctamente')
        return super().post(request, *args, **kwargs)

class bancosCreateView(LoginRequiredMixin,CreateView):
    model = t_banco
    form_class = bancoform
    template_name = 'bancos.html'
    context_object_name = 'bancos'
    success_url = reverse_lazy('bancos')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    
    def post(self, request, *args, **kwargs):
        # 👉 EXPORTAR EXCEL
        if 'exportar_excel' in request.POST:
            return self.exportar_excel()

        # 👉 SI VIENE EXCEL
        if 'archivo_excel' in request.FILES:
            form_excel = CargaExcelForm(request.POST, request.FILES)

            if form_excel.is_valid():
                archivo = request.FILES['archivo_excel']
                wb = load_workbook(archivo)
                ws = wb.active

                creados = 0
                omitidos = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    banco, codigo_ach, codigo_pse, codigo_br, codigo_fintech, estado = row

                    if not codigo_ach or not codigo_br or not codigo_pse or not codigo_fintech:
                        continue

                    _, created = t_banco.objects.get_or_create(
                        codigo_ach=codigo_ach,
                        codigo_pse=codigo_pse,
                        codigo_br=codigo_br,
                        codigo_fintech=codigo_fintech,
                        defaults={
                            'banco': banco,
                            'estado': True,
                            'user_creator': request.user,
                            'date_created': datetime.now()
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        omitidos += 1

                messages.success(
                    request,
                    f'Carga masiva exitosa. Nuevos: {creados}, Omitidos: {omitidos}'
                )

            return redirect(self.success_url)

        # 👉 SI NO, ES CREATE NORMAL
        return super().post(request, *args, **kwargs)

    def exportar_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bancos'

        ws.append(['banco', 'codigo_ach', 'codigo_pse', 'codigo_br', 'codigo_fintech'])

        registros = t_banco.objects.all().order_by('-pk')

        for r in registros:
            ws.append([r.banco, r.codigo_ach, r.codigo_pse, r.codigo_br, r.codigo_fintech])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=bancos_ss.xlsx'

        wb.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_banco.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        codigo = self.request.GET.get('codigo')
        
        banco = self.request.GET.get('banco')

        
        if banco:
            context['registros'] = t_banco.objects.filter(banco=banco)


        if codigo:
            context['registros'] = t_banco.objects.filter(codigo_pse=codigo)



        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            banco = t_banco.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(banco, field, value)
            banco.save()
            messages.success(self.request, 'Banco actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        
        return redirect(self.success_url)

    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class bancosDeleteView(LoginRequiredMixin,DeleteView):
    model = t_banco
    success_url = reverse_lazy('bancos')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Banco eliminado correctamente')
        return super().post(request, *args, **kwargs)


class entidadesssCreateView(LoginRequiredMixin,CreateView):
    model = t_entidadesss
    form_class = t_entidadesssform
    template_name = 'entidades_ss.html'
    context_object_name = 'entidades_ss'
    success_url = reverse_lazy('entidades_ss')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def post(self, request, *args, **kwargs):
        # 👉 EXPORTAR EXCEL
        if 'exportar_excel' in request.POST:
            return self.exportar_excel()

        # 👉 SI VIENE EXCEL
        if 'archivo_excel' in request.FILES:
            form_excel = CargaExcelForm(request.POST, request.FILES)

            if form_excel.is_valid():
                archivo = request.FILES['archivo_excel']
                wb = load_workbook(archivo)
                ws = wb.active

                creados = 0
                omitidos = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    tipo, codigo, nit, nombre = row

                    if not codigo:
                        continue

                    _, created = t_entidadesss.objects.get_or_create(
                        codigo=codigo,
                        defaults={
                            'tipo': tipo,
                            'nit': nit,
                            'nombre': nombre,
                            'user_creator': request.user,
                            'date_created': datetime.now()
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        omitidos += 1

                messages.success(
                    request,
                    f'Carga masiva exitosa. Nuevos: {creados}, Omitidos: {omitidos}'
                )

            return redirect(self.success_url)

        # 👉 SI NO, ES CREATE NORMAL
        return super().post(request, *args, **kwargs)

    def exportar_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Entidades SS'

        ws.append(['Tipo', 'Código', 'NIT', 'Nombre'])

        registros = t_entidadesss.objects.all().order_by('-pk')

        for r in registros:
            ws.append([r.tipo, r.codigo, r.nit, r.nombre])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=entidades_ss.xlsx'

        wb.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_entidadesss.objects.all().order_by('-pk')[:30]
        context['registros'] = registros
        
        codigo = self.request.GET.get('codigo')
        
        tipo = self.request.GET.get('tipo')

        if tipo:
            context['registros'] = t_entidadesss.objects.filter(tipo=tipo)

        if codigo:
            context['registros'] = t_entidadesss.objects.filter(codigo=codigo)


        context['form_excel'] = CargaExcelForm()
        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            entidad = t_entidadesss.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(entidad, field, value)
            entidad.save()
            messages.success(self.request, 'Entidad actualizada correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.success_url)

    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class entidadesssDeleteView(LoginRequiredMixin,DeleteView):
    model = t_entidadesss
    success_url = reverse_lazy('entidades_ss')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Banco eliminado correctamente')
        return super().post(request, *args, **kwargs)


class t_conceptos_salarioCreateView(LoginRequiredMixin,CreateView):
    model = t_conceptos_salario
    form_class = t_conceptos_salarioform
    template_name = 'conceptos_salario.html'
    context_object_name = 'conceptos_salario'
    success_url = reverse_lazy('conceptos_salario')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_conceptos_salario.objects.all().order_by('-pk')[:30]
        context['registros'] = registros
        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            conceptossalario = t_conceptos_salario.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(conceptossalario, field, value)
            conceptossalario.save()
            messages.success(self.request, 'Concepto salario actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.success_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Validar campos del formulario %s'%form.errors)
        return super().form_invalid(form)
    
class t_conceptos_salarioDeleteView(LoginRequiredMixin,DeleteView):
    model = t_conceptos_salario
    success_url = reverse_lazy('conceptos_salario')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de contraro eliminado correctamente')
        return super().post(request, *args, **kwargs)
    
    
class tipo_nominaCreateView(LoginRequiredMixin,CreateView):
    model = t_tipo_nomina
    form_class = t_tipo_nominaform
    template_name = 'tipo_nomina.html'
    context_object_name = 'tipo_nomina'
    success_url = reverse_lazy('tipo_nomina')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = t_tipo_nomina.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        codigo = self.request.GET.get('codigo')
        
        estado = self.request.GET.get('estado')

        
        if codigo:
            context['registros'] = t_tipo_nomina.objects.filter(codigo=codigo)

        
        if estado:
            context['registros'] = t_tipo_nomina.objects.filter(estado=estado)


        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            tipo_nomina = t_tipo_nomina.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(tipo_nomina, field, value)
            tipo_nomina.save()
            messages.success(self.request, 'Tipo de nomina actualizada correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)

        return redirect(self.success_url)
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class tipo_nominaDeleteView(LoginRequiredMixin,DeleteView):
    model = t_tipo_nomina
    success_url = reverse_lazy('tipo_nomina')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de nomina eliminada correctamente')
        return super().post(request, *args, **kwargs)


    
class parametro_generalCreateView(LoginRequiredMixin,CreateView):
    model = ParametroGeneral
    form_class = t_parametrogeneralform
    template_name = 'parametro_general.html'
    context_object_name = 'parametro_general'
    success_url = reverse_lazy('parametro_general')
    login_url = '/accounts/login/'           # opcional
    redirect_field_name = 'next'  # opcional

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        pk = self.request.POST.get('pk')

        if pk:
            kwargs['instance'] = ParametroGeneral.objects.get(pk=pk)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = ParametroGeneral.objects.all().order_by('-pk')[:30]
        context['registros'] = registros

        codigo = self.request.GET.get('codigo')
        
        estado = self.request.GET.get('estado')

        
        if codigo:
            context['registros'] = ParametroGeneral.objects.filter(codigo=codigo)

        
        if estado:
            context['registros'] = ParametroGeneral.objects.filter(activo=estado)


        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if  pk:
            parametro = ParametroGeneral.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(parametro, field, value)
            parametro.save()
            messages.success(self.request, 'Parametro actualizado correctamente')
        else:
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        
        return redirect(self.success_url)
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
class parametro_generalDeleteView(LoginRequiredMixin,DeleteView):
    model = ParametroGeneral
    success_url = reverse_lazy('parametro_general')
    login_url = 'accounts/login'

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Parametro eliminado correctamente')
        return super().post(request, *args, **kwargs)


class parametro_DetalleCreateView(LoginRequiredMixin,CreateView):
    model = ParametroDetalle
    form_class = t_parametrodetllleform
    template_name = 'parametro_detalle.html'
  
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        pk = self.request.POST.get('pk')

        if pk:
            kwargs['instance'] = ParametroDetalle.objects.get(pk=pk)
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        parametro = get_object_or_404(ParametroGeneral,pk=self.kwargs['parametro_id'])

        context['parametro'] = parametro
        context['registros'] = ParametroDetalle.objects.filter(parametro=parametro)

        return context
     
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            parametro = ParametroDetalle.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(parametro, field, value)
            parametro.save()
            messages.success(self.request, 'Parametro actualizado correctamente')
        else:
            form.instance.parametro_id = self.kwargs['parametro_id']
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())

    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'parametro_detalle',
            kwargs={'parametro_id': self.kwargs['parametro_id']}
        )


class parametro_detalleDeleteView(LoginRequiredMixin,DeleteView):
    model = ParametroDetalle
    login_url = 'accounts/login'

    def get_success_url(self):
        return reverse_lazy(
            'parametro_detalle',
            kwargs={'parametro_id': self.object.parametro_id}
        )
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)
