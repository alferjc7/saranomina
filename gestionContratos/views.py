from datetime import datetime
from operator import concat
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import CreateView, DeleteView
from openpyxl import Workbook
from gestionContratos.models import (t_contrato, t_contrato_banco, 
                                     t_contrato_entidadesss, t_contrato_salario,
                                     t_contrato_deducibles)
from gestionContratos.forms import (t_contratoform, t_contrato_banco_form,
                                    t_contrato_entidadesss_form, t_contrato_salario_form,
                                    t_contrato_deducible_form)
from parametros.models import t_subtipo_cotizante, t_entidadesss
from django.http import HttpResponse, JsonResponse
from openpyxl.utils import get_column_letter



# Create your views here.
def cargar_subtipos(request):
    tipo_id = request.GET.get('tipo_id')
    subtipos = t_subtipo_cotizante.objects.filter(
        codigo_cotizante=tipo_id
    ).values('id', 'descripcion')
    return JsonResponse(list(subtipos), safe=False)

def ajax_entidades_por_tipo(request):
    tipo = request.GET.get('tipo')
    entidades = t_entidadesss.objects.filter(tipo=tipo)

    data = [{'id': e.id, 'nombre': e.nombre} for e in entidades]
    return JsonResponse(data, safe=False)

class t_contratosCreateView(LoginRequiredMixin,CreateView):
    model = t_contrato
    form_class = t_contratoform
    template_name = 'contratos.html'
    context_object_name = 'contratos'
    success_url = reverse_lazy('contratos')
    login_url = '/accounts/login/'          
    redirect_field_name = 'next'  

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Contratos'

            ws.append(['EMPRESA', 'IDENTIFICACION', 'NOMBRE COMPLETO', 'CODIGO CONTRATO', 'FECHA INGRESO' ,
                       'FECHA FIN','TIPO CONTRATO','PERIODO VACACIONES', 'TIPO COTIZANTE', 'SUBTIPO COTIZANTE',
                       'CODIGO INTERNO','PROCEDIMIENTO RET','PORCENTAJE RET','ESTADO','REGIMEN CESANTIAS'])

            registros = t_contrato.objects.filter(empresa_id = self.request.session.get('empresa_id'))

            for r in registros:
                ws.append([r.empresa.razon_social, r.identificacion.identificacion, 
                           f"{r.identificacion.nombre or ''} {r.identificacion.segundo_nombre or ''} {r.identificacion.apellido or ''} {r.identificacion.segundo_apellido or ''}".strip(),
                           r.cod_contrato, r.fecha_ingreso, r.fecha_fin, r.tipo_contrato.contrato, r.periodo_vac,
                           r.tipo_cotizante.descripcion, r.subtipo_cotizante.descripcion, r.codigo_interno, r.get_procedimiento_display(),
                           r.porcentaje_retencion, r.get_estado_display(), r.get_cesantias_display()])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Contratos.xlsx'
            wb.save(response)
            return response 

        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        identificacion = self.request.GET.get('identificacion')
        contrato = self.request.GET.get('contrato')

        registros = t_contrato.objects.filter(empresa_id = self.request.session.get('empresa_id'))
        
        if identificacion:
            registros = t_contrato.objects.filter(empresa_id = self.request.session.get('empresa_id'), identificacion__identificacion = identificacion)

        if contrato:
            registros = t_contrato.objects.filter(empresa_id = self.request.session.get('empresa_id'),cod_contrato = contrato)
            
        context['registros'] = registros
        return context
    
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            contrato = t_contrato.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(contrato, field, value)
            contrato.save()
            messages.success(self.request, 'Contrato actualizado correctamente')
        else:
            form.instance.empresa_id = self.request.session.get('empresa_id')
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        
        return redirect(self.success_url)

    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)

class t_contrato_bancoCreateView(LoginRequiredMixin,CreateView):
    model = t_contrato_banco
    form_class = t_contrato_banco_form
    template_name = 'contrato_banco.html'
    
    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Contrato_banco'

            ws.append(['EMPRESA', 'CONTRATO', 'NOMBRE COMPLETO', 'BANCO', 'TIPO CUENTA', 'NUMERO CUENTA', 
                       'FECHA INICIO' ,'FECHA FIN','ESTADO'])

            contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        
            registros = t_contrato_banco.objects.filter(contrato=contrato)

            for r in registros:
                ws.append([r.contrato.empresa.razon_social, r.contrato.cod_contrato, 
                           f"{r.contrato.identificacion.nombre or ''} {r.contrato.identificacion.segundo_nombre or ''} {r.contrato.identificacion.apellido or ''} {r.contrato.identificacion.segundo_apellido or ''}".strip(),
                           r.banco.banco, r.get_cuenta_display() , r.numero_cuenta, r.fecha_inicio, r.fecha_fin, 
                           'ACTIVO' if r.estado else 'INACTIVO'])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Contrato_banco.xlsx'
            wb.save(response)
            return response 
        return super().post(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])

        context['contrato'] = contrato
        context['registros'] = t_contrato_banco.objects.filter(contrato=contrato)

        estado = self.request.GET.get('estado')

        if estado:
            context['registros'] = t_contrato_banco.objects.filter(contrato=contrato, estado=estado)

        return context
     
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            contratoban = t_contrato_banco.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(contratoban, field, value)
            contratoban.save()
            messages.success(self.request, 'Contrato banco actualizado correctamente')
        else:
            form.instance.contrato_id = self.kwargs['contrato_id']
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'contrato_banco',
            kwargs={'contrato_id': self.kwargs['contrato_id']}
        )


class t_contrato_bancoDeleteView(LoginRequiredMixin,DeleteView):
    model = t_contrato_banco
    login_url = 'accounts/login'

    def get_success_url(self):
        return reverse_lazy(
            'contrato_banco',
            kwargs={'contrato_id': self.object.contrato_id}
        )
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)


class t_contrato_entidadCreateView(LoginRequiredMixin,CreateView):
    model = t_contrato_entidadesss
    form_class = t_contrato_entidadesss_form
    template_name = 'contrato_entidades_ss.html'
    
    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Contrato_entidades'

            ws.append(['EMPRESA', 'CONTRATO', 'NOMBRE COMPLETO', 'TIPO ENTIDAD','CODIGO ENTIDAD' ,'ENTIDAD', 
                       'FECHA INICIO' ,'FECHA FIN'])

            contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        
            registros = t_contrato_entidadesss.objects.filter(contrato=contrato)

            for r in registros:
                ws.append([r.contrato.empresa.razon_social, r.contrato.cod_contrato, 
                           f"{r.contrato.identificacion.nombre or ''} {r.contrato.identificacion.segundo_nombre or ''} {r.contrato.identificacion.apellido or ''} {r.contrato.identificacion.segundo_apellido or ''}".strip(),
                           r.get_tipo_entidad_display(), r.entidad.codigo, r.entidad.nombre, r.fecha_inicio, r.fecha_fin])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Contrato_entidades.xlsx'
            wb.save(response)
            return response 
        
        return super().post(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])

        context['contrato'] = contrato
        context['registros'] = t_contrato_entidadesss.objects.filter(contrato=contrato)
        
        tipo = self.request.GET.get('tipo')

        if tipo:
            context['registros'] = t_contrato_entidadesss.objects.filter(contrato=contrato,tipo_entidad = tipo)

        return context
     
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            contratoent = t_contrato_entidadesss.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(contratoent, field, value)
            contratoent.save()
            messages.success(self.request, 'Contrato entidad actualizada correctamente')
        else:
            form.instance.contrato_id = self.kwargs['contrato_id']
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'contrato_entidad_ss',
            kwargs={'contrato_id': self.kwargs['contrato_id']}
        )


class t_contrato_entidadDeleteView(LoginRequiredMixin,DeleteView):
    model = t_contrato_entidadesss
    login_url = 'accounts/login'

    def get_success_url(self):
        return reverse_lazy(
            'contrato_entidad_ss',
            kwargs={'contrato_id': self.object.contrato_id}
        )
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)


class t_contrato_salarioCreateView(LoginRequiredMixin,CreateView):
    model = t_contrato_salario
    form_class = t_contrato_salario_form
    template_name = 'contrato_salario.html'

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Contrato_salario'

            ws.append(['EMPRESA', 'CONTRATO', 'NOMBRE COMPLETO', 'TIPO SALARIO', 'FECHA INICIO' ,
                       'FECHA FIN','ESTADO','RETROACTIVO'])

            contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        
            registros = t_contrato_salario.objects.filter(contrato=contrato)

            for r in registros:
                ws.append([r.contrato.empresa.razon_social, r.contrato.cod_contrato, 
                           f"{r.contrato.identificacion.nombre or ''} {r.contrato.identificacion.segundo_nombre or ''} {r.contrato.identificacion.apellido or ''} {r.contrato.identificacion.segundo_apellido or ''}".strip(),
                           r.tipo_salario.salario, r.fecha_inicio, r.fecha_fin, 'ACTIVO' if r.estado else 'INACTIVO', 'SI' if r.retroactivo else 'NO'])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Contrato_salario.xlsx'
            wb.save(response)
            return response 

        return super().post(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        
        context['contrato'] = contrato
        context['registros'] = t_contrato_salario.objects.filter(contrato=contrato)
        estado = self.request.GET.get('estado')

        if estado:
            context['registros'] = t_contrato_salario.objects.filter(contrato=contrato,estado=estado)

        return context
     
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            contratosal = t_contrato_salario.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(contratosal, field, value)
            contratosal.save()
            messages.success(self.request, 'Salario contrato actualizado correctamente')
        else:
            form.instance.contrato_id = self.kwargs['contrato_id']
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'contrato_salario',
            kwargs={'contrato_id': self.kwargs['contrato_id']}
        )


class t_contrato_salarioDeleteView(LoginRequiredMixin,DeleteView):
    model = t_contrato_salario
    login_url = 'accounts/login'

    def get_success_url(self):
        return reverse_lazy(
            'contrato_salario',
            kwargs={'contrato_id': self.object.contrato_id}
        )
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)


class t_contrato_deducibleCreateView(LoginRequiredMixin,CreateView):
    model = t_contrato_deducibles
    form_class = t_contrato_deducible_form
    template_name = 'contrato_deducible.html'

    def post(self, request, *args, **kwargs):
        accion = request.POST.get('accion')
        
        if accion == 'exportar_excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Contrato_deducibles'

            ws.append(['EMPRESA', 'CONTRATO', 'NOMBRE COMPLETO', 'TIPO DEDUCIBLE', 'VALOR', 'FECHA INICIO' ,
                       'FECHA FIN'])

            contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        
            registros = t_contrato_deducibles.objects.filter(contrato=contrato)

            for r in registros:
                ws.append([r.contrato.empresa.razon_social, r.contrato.cod_contrato, 
                           f"{r.contrato.identificacion.nombre or ''} {r.contrato.identificacion.segundo_nombre or ''} {r.contrato.identificacion.apellido or ''} {r.contrato.identificacion.segundo_apellido or ''}".strip(),
                           r.get_tipo_deducible_display(), r.valor, r.fecha_inicio, r.fecha_fin])

            for column_cells in ws.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column].width = max_length + 2 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' )
            response['Content-Disposition'] = 'attachment; filename=Contrato_deducibles.xlsx'
            wb.save(response)
            return response 

        return super().post(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contrato'] = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])
        return kwargs  
  
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contrato = get_object_or_404(t_contrato,pk=self.kwargs['contrato_id'])

        context['contrato'] = contrato
        context['registros'] = t_contrato_deducibles.objects.filter(contrato=contrato)

        tipo = self.request.GET.get('tipo')

        if tipo:
            context['registros'] = t_contrato_deducibles.objects.filter(contrato=contrato,tipo_deducible=tipo)


        return context
     
    def form_valid(self, form):
        pk = self.request.POST.get('pk')
        if pk:
            contratoded = t_contrato_deducibles.objects.get(pk=pk)
            for field, value in form.cleaned_data.items():
                setattr(contratoded, field, value)
            contratoded.save()
            messages.success(self.request, 'Contrato deducible actualizado correctamente')
        else:
            form.instance.contrato_id = self.kwargs['contrato_id']
            form.instance.date_created = datetime.now()
            form.instance.user_creator = self.request.user
            messages.success(self.request, 'Registro creado correctamente')
            return super().form_valid(form)
        return redirect(self.get_success_url())

    
    def form_invalid(self, form):
        print(form.errors)
        messages.error(self.request, 'Validar campos del formulario')
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse(
            'contrato_deducible',
            kwargs={'contrato_id': self.kwargs['contrato_id']}
        )


class t_contrato_deducibleDeleteView(LoginRequiredMixin,DeleteView):
    model = t_contrato_deducibles
    login_url = 'accounts/login'

    def get_success_url(self):
        return reverse_lazy(
            'contrato_deducible',
            kwargs={'contrato_id': self.object.contrato_id}
        )
    def post(self, request, *args, **kwargs):
        messages.success(request, 'Registro eliminado correctamente')
        return super().post(request, *args, **kwargs)
